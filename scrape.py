import requests
from bs4 import BeautifulSoup
import re
import time
import csv 
import sys
import psycopg2
import openpyxl
import os 

DB_USER_NAME = os.environ['DB_USER_NAME']
DB_PASSWORD = os.environ['DB_PASSWORD']
DB_DATABASE = os.environ['DB_DATABASE']
DB_HOST = os.environ['DB_HOST']

def get_soup(url, write_file=False):
    #Return a BeautifulSoup4 if request to url is successful
    #write_file = True writes the html file to disk

    #Create a request, set headers to bypass basic carwale.com scraping security
    page = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'})
    #Raise exception if invalid http status code returned 
    page.raise_for_status()
    #Create BeautifulSoup4 object using lxml parser
    soup = BeautifulSoup(page.text, "lxml")

    if write_file:
        file = open('page.html', 'w')
        file.write(page.text)
        file.close()

    return soup

def write_csv(name, company, image, summary, price_starting, price_topend, mileage_l, mileage_u, manual, automatic, petrol, diesel, cng, electric, seating):
    #Writes scraped car data to a csv file
    file = open('data.csv', 'a')
    #Create csv.write object
    csvWriter = csv.writer(file)
    #Write the row of values to the csv file 
    csvWriter.writerow([name, company, image, summary, price_starting, price_topend, mileage_l, mileage_u, manual, automatic, petrol, diesel, cng, electric, seating])
    file.close()


def write_db(dbcon, name, company, image, summary, price_starting, price_topend, mileage_l, mileage_u, manual, automatic, petrol, diesel, cng, electric, seating):
    #Writes scraped car data to the databse

    #create databse cursor
    cur = dbcon.cursor()

    #Sql query for insert
    SQL = '''INSERT INTO cars(name, company, image, summary, price_starting, price_topend, mileage_l, mileage_u, manual, automatic, petrol, diesel, cng, electric, seating_capacity)
            VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'''
    
    #params to be substituted into sql string
    params = (name, company, image, summary, price_starting, price_topend, mileage_l, mileage_u, manual, automatic, petrol, diesel, cng, electric, seating)

    #execute sql query
    cur.execute(SQL, params)

    #close databse cursor
    cur.close()

    #commit databse transaction
    dbcon.commit()

def write_excel(name, company, image, summary, price_starting, price_topend, mileage_l, mileage_u, manual, automatic, petrol, diesel, cng, electric, seating):
    #writes scraped data to excel file 
    try:
        #Open Excel workbook
        wb = openpyxl.load_workbook('cars.xlsx')
        #Get active sheet
        sheet=wb.active
        #Add row of values to sheet
        sheet.append([name, company, image, summary, price_starting, price_topend, mileage_l, mileage_u, manual, automatic, petrol, diesel, cng, electric, seating])
        #Save Excel file
        wb.save('cars.xlsx')
    except (FileNotFoundError, openpyxl.exceptions.InvalidFileException):
        #Excel file doesnt exist 
        wb=openpyxl.Workbook()
        sheet= wb.active
        #Set Sheet title
        sheet.title='CAR DETAILS'
        #Add column names
        column_names=['NAME', 'COMPANY', 'IMAGE', 'SUMMARY', 'PRICE_STARTING', 'PRICE_TOPEND', 'MILEAGE_L', 'MILEAGE_U', 'MANUAL', 'AUTOMATIC', 'PETROL', 'DIESEL', 'CNG', 'ELECTRIC', 'SEATING']        
        for i in range(len(column_names)):
            sheet.cell(row=1,column=i+1).value=column_names[i]
        sheet.append([name, company, image, summary, price_starting, price_topend, mileage_l, mileage_u, manual, automatic, petrol, diesel, cng, electric, seating])
        wb.save('cars.xlsx')

def log_done(company):
    #Lop if scrape was succesful
    file = open('log.txt', 'a')
    file.write('\nScraped: ' + company)
    file.close()

def scrape_car_company(dbcon, company_url, company, output):
    domain = 'https://www.carwale.com'

    print('Scraping ' + company)

    soup = get_soup(company_url)

    car_urls = []

    #Find url to pages of all cars of a company
    li_tags = soup.find_all('li', class_='o-fzptUA')
    for li_tag in li_tags: 
        #Skip upcoming cars
        if li_tag.find(class_='o-fzoTov'):
            print('Skipping: ' + li_tag.find('a', class_='o-fzoHMp')['href'])
            continue

        if li_tag.find('a', class_='o-fzoHMp'):
            car_url = domain + li_tag.find('a', class_='o-fzoHMp')['href']
        elif li_tag.find('a', class_='o-fzpilz'):
            car_url = domain + li_tag.find('a', class_='o-fzpilz')['href']

        car_urls.append(car_url)
        print(car_url)
  
    #Scrape a particular cars page
    for car_url in car_urls:    
        print('Scraping ' + car_url)
        soup = get_soup(car_url, write_file=False)

        #Get car's name
        name = soup.find('h1', class_="o-eqqVmt").find(text=True, recursive=False)

        #Get car's image url 
        image = soup.find('img', class_="o-bXKmQE")['src']

        #Get a summary of car's details
        summary = soup.find('div', class_="o-fyWCgU").text.strip()
        
        #Get the price range of the car
        priceRegex = re.compile(r'[0-9]+\.[0-9]+')

        price = priceRegex.findall(soup.find('div', class_='o-fyWCgU').text)
        price_starting = float(price[0])
        price_topend = float(price[1])

        table = soup.find('table',class_='o-bfyaNx')

        #Get the mileage of the car
        mileageHeader = table.find('span', text='Mileage')
        if mileageHeader:
            mileageRegex = re.compile(r'[0-9]+\.[0-9]+|[0-9]+')

            mileage = mileageRegex.findall(mileageHeader.parent.next_sibling.text)

            if len(mileage) == 1:
                mileage_l = float(mileage[0])
                mileage_u = float(mileage[0])
            else:
                mileage_l = float(mileage[0])
                mileage_u = float(mileage[1])
        else:
            mileage_l = 17.0
            mileage_u = 21.0

        #Get the transmission type of the car: Petrol, Automatic
        transmissionHeader = table.find('span', text='Transmission')
        if transmissionHeader:
            automatic = False
            manual = False


            transmission_str = transmissionHeader.parent.next_sibling.text.lower()

            if 'automatic' in transmission_str or 'amt' in transmission_str:
                automatic = True

            if 'manual' in transmission_str:
                manual = True
        else:
            automatic = True
            manual = True

        #Get the fuel type of the car: Petrol, Diesel, CNG, Electric
        fuelHeader = table.find('span', text='Fuel Type')
        if fuelHeader:
            fuel_str = fuelHeader.parent.next_sibling.text.lower()

            petrol = False
            diesel = False
            cng = False
            electric = False

            if 'petrol' in fuel_str:
                petrol = True

            if 'diesel' in fuel_str:
                diesel = True

            if 'cng' in fuel_str:
                cng = True

            if 'electric' in fuel_str:
                electric = True
        else:
            petrol = True
            diesel = True
            cng = False
            electric = False

        #Get seating capacity of car
        seatingHeader = table.find('span', text='Seating Capacity')
        if seatingHeader:
            seatingRegex = re.compile('[0-9]+')

            seating = int(seatingRegex.findall(seatingHeader.parent.next_sibling.text)[-1])
        else:
            seating = 5

        if output == 1:
            #write scraped details to database
            write_db(dbcon, name, company, image, summary, price_starting, price_topend, mileage_l, mileage_u, manual, automatic, petrol, diesel, cng, electric, seating)
        elif output == 2:
            #write scraped details to csv file
            write_csv(name, company, image, summary, price_starting, price_topend, mileage_l, mileage_u, manual, automatic, petrol, diesel, cng, electric, seating)
        else:
            #write scraped details to excel file
            write_excel(name, company, image, summary, price_starting, price_topend, mileage_l, mileage_u, manual, automatic, petrol, diesel, cng, electric, seating)

        #add delay of 5 seconds between requests
        time.sleep(5)

    #log if scrape successful
    log_done(company)

if __name__ == "__main__":
    print("Carwale.com Scraper: ")

    #Get company url
    company_url = input('Enter car company url: ')
    #Get company name
    company = input('Enter company name: ')

    #Get where to scrape data to
    print('Enter where you want to scrape data to:\n1.Database\n2.CSV File\n3.Excel File')
    output = int(input('Enter choice: '))

    if output in [2,3]:
        dbcon = None
    elif output == 1:
        #Database connection string
            dbcon = psycopg2.connect(
            user=DB_USER_NAME,
            password=DB_PASSWORD,
            database=DB_DATABASE,
            host=DB_HOST)  
    else:
        print('Invalid choice')
        sys.exit()

    #Scrape a company's car details
    scrape_car_company(dbcon, company_url, company, output)

    #close database connection
    if output == 1:
        dbcon.close()
